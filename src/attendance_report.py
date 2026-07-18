"""
Attendance Report Writer for IR Attendance client.
Handles:
- Writing opening/closing events to local JSON files
- Converting JSON → XLSX
- Pushing XLSX files to Google Drive at midnight (background thread)
"""
import os
import json
import threading
import time
from pathlib import Path
from datetime import datetime, date


REPORTS_DIR = Path("C:/IR Attendance/Reports")


def _ensure_reports_dir():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _log_upload_event(message):
    try:
        _ensure_reports_dir()
        log_path = REPORTS_DIR / "upload_log.txt"
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {message}\n")
    except Exception:
        pass


def _get_json_path(batch_id, date_str):
    """Return local JSON path for a batch + date."""
    return REPORTS_DIR / batch_id / f"{date_str}.json"


def _read_json(batch_id, date_str):
    """Load existing JSON report or return empty structure."""
    path = _get_json_path(batch_id, date_str)
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "batch_id": batch_id,
        "batch_name": "",
        "date": date_str,
        "in_time": "",
        "out_time": "",
        "total_students": 0,
        "entries": {}
    }


def _write_json(data, batch_id, date_str):
    """Save JSON report file."""
    _ensure_reports_dir()
    path = _get_json_path(batch_id, date_str)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def log_attendance_event(attendance_id, batch_id, batch_name,
                         event_type, timestamp, in_time="", out_time="", total_students=0, username="default"):
    """
    Write an opening or closing event to the local daily JSON report.
    event_type: "opening" or "closing"
    timestamp:  "HH:MM:SS" string
    Thread-safe using a per-batch lock.
    """
    _ensure_reports_dir()
    date_str = date.today().isoformat()
    data = _read_json(batch_id, date_str)

    # Update batch-level metadata if provided
    if batch_name:
        data["batch_name"] = batch_name
    if in_time:
        data["in_time"] = in_time
    if out_time:
        data["out_time"] = out_time
    if total_students:
        data["total_students"] = total_students

    # Update entry for this attendance ID
    entry = data["entries"].get(attendance_id, {"attendance_id": attendance_id})
    if event_type == "opening":
        entry["opening_time"] = timestamp
        # Only set closing_time key if not already present
        if "closing_time" not in entry:
            entry["closing_time"] = None
    elif event_type == "closing":
        entry["closing_time"] = timestamp
        if "opening_time" not in entry:
            entry["opening_time"] = None

    data["entries"][attendance_id] = entry
    _write_json(data, batch_id, date_str)
    
    # Trigger real-time upload to Google Drive
    trigger_realtime_push(batch_id, date_str, username)


def trigger_realtime_push(batch_id, date_str, username):
    """Start a background thread to upload the daily JSON to Google Drive."""
    threading.Thread(
        target=_realtime_push_worker,
        args=(batch_id, date_str, username),
        daemon=True
    ).start()


def _realtime_push_worker(batch_id, date_str, username):
    try:
        from drive_manager import drive_manager
        if not drive_manager.is_authenticated():
            _log_upload_event(f"[RealtimePush] Drive not authenticated, skipping push for {batch_id}/{date_str}")
            print("[RealtimePush] Drive not authenticated, skipping push")
            return
        
        json_path = _get_json_path(batch_id, date_str)
        if not json_path.exists():
            print(f"[RealtimePush] JSON file {json_path} does not exist")
            return
            
        print(f"[RealtimePush] Uploading {batch_id}/{date_str}.json ...")
        ok, msg = drive_manager.upload_json(str(json_path), username, date_str, batch_id)
        
        log_msg = f"[RealtimePush] Upload {batch_id}/{date_str}.json (User: {username}) -> {'SUCCESS' if ok else 'FAIL: ' + str(msg)}"
        if drive_manager.is_service_account and not ok:
            log_msg += " (Warning: Service Accounts have 0-byte quota unless sharing a Shared Drive. Use OAuth 2.0 or check setup.)"
        _log_upload_event(log_msg)
        print(f"[RealtimePush] {'OK' if ok else 'FAIL'}: {msg}")
    except Exception as e:
        _log_upload_event(f"[RealtimePush] Critical Error: {e}")
        print(f"[RealtimePush] Error: {e}")


def json_to_xlsx(batch_id, date_str, json_data=None):
    """
    Convert a daily JSON report to an XLSX file.
    Returns path to the XLSX file.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if json_data is None:
        json_data = _read_json(batch_id, date_str)

    xlsx_path = REPORTS_DIR / batch_id / f"{date_str}.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Styles
    bold14   = Font(bold=True, size=14)
    bold11   = Font(bold=True, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    thin     = Side(style="thin")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    pre_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    abs_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"IR Attendance Report — Batch {batch_id}"
    ws["A1"].font = bold14
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 24

    # Row 2: Meta info
    total = json_data.get("total_students", len(json_data.get("entries", {})))
    meta = [
        f"Batch ID: {json_data.get('batch_id', batch_id)}",
        f"Date: {date_str}",
        f"In Time: {json_data.get('in_time', '—')}",
        f"Out Time: {json_data.get('out_time', '—')}",
        f"Total Students: {total}"
    ]
    for col_idx, text in enumerate(meta, 1):
        c = ws.cell(row=2, column=col_idx, value=text)
        c.font = bold11
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # Row 4: Column headers
    col_headers = ["Sr No", "Attendance ID", "Opening Time", "Closing Time", "Status"]
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 18

    # Data rows (sorted by attendance_id)
    entries = sorted(json_data.get("entries", {}).values(), key=lambda e: str(e.get("attendance_id", "")))
    for sr, entry in enumerate(entries, 1):
        opening = entry.get("opening_time") or "—"
        closing = entry.get("closing_time") or "—"
        status = "Present" if entry.get("opening_time") else "Absent"
        fill = pre_fill if status == "Present" else abs_fill
        row_data = [sr, entry.get("attendance_id", ""), opening, closing, status]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=sr + 4, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = center
            cell.border = border

    for col, width in zip("ABCDE", [10, 20, 16, 16, 12]):
        ws.column_dimensions[col].width = width

    wb.save(str(xlsx_path))
    return str(xlsx_path)


# ──────────────────────────────────────────────────────────────────────────────
#  Midnight push background thread
# ──────────────────────────────────────────────────────────────────────────────

_push_lock = threading.Lock()
_push_thread = None
_stop_push = False


def start_midnight_push_worker(username, get_drive_manager_fn):
    """
    Start the background midnight push thread.
    username: the licensed username (used for Drive folder path)
    get_drive_manager_fn: callable that returns the DriveManager instance
    """
    global _push_thread, _stop_push
    _stop_push = False
    _push_thread = threading.Thread(
        target=_midnight_push_loop,
        args=(username, get_drive_manager_fn),
        daemon=True
    )
    _push_thread.start()


def stop_midnight_push_worker():
    global _stop_push
    _stop_push = True


def _midnight_push_loop(username, get_drive_manager_fn):
    """
    Runs forever (daemon thread).
    Watches for day change — at midnight, converts yesterday's JSONs to XLSX and pushes to Drive.
    Also retries any pending pushes from previous sessions.
    """
    _ensure_reports_dir()
    last_date = date.today().isoformat()
    pending_flag = REPORTS_DIR / ".pending_push"

    # On startup: retry any pending push from last session
    if pending_flag.exists():
        try:
            pending_date = pending_flag.read_text().strip()
            if pending_date and pending_date != last_date:
                _push_date(pending_date, username, get_drive_manager_fn)
                pending_flag.unlink(missing_ok=True)
        except Exception as e:
            print(f"[MidnightPush] Startup retry failed: {e}")

    while not _stop_push:
        time.sleep(10)  # check every 10 seconds (very lightweight)
        if _stop_push:
            break

        today = date.today().isoformat()
        if today != last_date:
            # Day has changed — push yesterday's data
            yesterday = last_date
            # Write pending flag (in case machine goes offline mid-push)
            try:
                pending_flag.write_text(yesterday)
            except Exception:
                pass

            _push_date(yesterday, username, get_drive_manager_fn)

            try:
                pending_flag.unlink(missing_ok=True)
            except Exception:
                pass

            last_date = today


def _push_date(date_str, username, get_drive_manager_fn):
    """Upload all batch JSONs for a given date to Drive."""
    try:
        dm = get_drive_manager_fn()
        if not dm or not dm.is_authenticated():
            _log_upload_event(f"[MidnightPush] Drive not authenticated, skipping push for {date_str}")
            print(f"[MidnightPush] Drive not authenticated, skipping push for {date_str}")
            return

        # Scan all batch folders
        if not REPORTS_DIR.exists():
            return

        for batch_folder in REPORTS_DIR.iterdir():
            if not batch_folder.is_dir():
                continue
            batch_id = batch_folder.name
            json_path = batch_folder / f"{date_str}.json"
            if not json_path.exists():
                continue

            try:
                print(f"[MidnightPush] Uploading {batch_id}/{date_str}.json ...")
                ok, msg = dm.upload_json(str(json_path), username, date_str, batch_id)
                log_msg = f"[MidnightPush] Upload {batch_id}/{date_str}.json (User: {username}) -> {'SUCCESS' if ok else 'FAIL: ' + str(msg)}"
                if dm.is_service_account and not ok:
                    log_msg += " (Warning: Service Accounts have 0-byte quota unless sharing a Shared Drive. Use OAuth 2.0 or check setup.)"
                _log_upload_event(log_msg)
                print(f"[MidnightPush] {'OK' if ok else 'FAIL'}: {msg}")

            except Exception as e:
                _log_upload_event(f"[MidnightPush] Error processing {batch_id}/{date_str}: {e}")
                print(f"[MidnightPush] Error processing {batch_id}/{date_str}: {e}")

    except Exception as e:
        _log_upload_event(f"[MidnightPush] Push loop critical error: {e}")
        print(f"[MidnightPush] Push loop error: {e}")
