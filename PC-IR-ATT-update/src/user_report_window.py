"""
User Report Windows for IR Attendance Admin Panel.
- UserReportWindow: shows batch summary for a user
- BatchDetailWindow: calendar + daily sheet + XLSX download
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import json
from pathlib import Path
from datetime import datetime, date
import threading


def _read_xlsx_to_rows(xlsx_path):
    """Read an XLSX file and return (header_info dict, list of row dicts)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}, []

        # Parse header block (first 3 rows)
        header_info = {}
        data_rows = []
        col_headers = []
        data_start = 0

        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).startswith("Batch ID"):
                # Row: Batch ID: 12345  |  Date: ...  |  In Time: ...  |  Out Time: ...  | Total: ...
                for cell in row:
                    if cell:
                        s = str(cell)
                        for key in ["Batch ID", "Date", "In Time", "Out Time", "Total Students"]:
                            if s.startswith(key + ":"):
                                header_info[key] = s.split(":", 1)[1].strip()
            elif row and row[0] and str(row[0]).strip() in ("Sr No", "Sr"):
                col_headers = [str(c).strip() if c else "" for c in row]
                data_start = i + 1
                break

        # Parse data rows
        for row in rows[data_start:]:
            if not any(row):
                continue
            entry = {}
            for j, val in enumerate(row):
                if j < len(col_headers):
                    entry[col_headers[j]] = str(val).strip() if val is not None else ""
            data_rows.append(entry)

        wb.close()
        return header_info, data_rows
    except Exception as e:
        print(f"XLSX read error: {e}")
        return {}, []


def _read_json_to_rows(json_path):
    """Read a daily JSON report and return (header_info dict, list of row dicts)."""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        header_info = {
            "Batch ID": data.get("batch_id", ""),
            "Date": data.get("date", ""),
            "In Time": data.get("in_time", ""),
            "Out Time": data.get("out_time", ""),
            "Total Students": str(data.get("total_students", 0))
        }

        data_rows = []
        entries = sorted(data.get("entries", {}).values(), key=lambda e: str(e.get("attendance_id", "")))
        for sr, entry in enumerate(entries, 1):
            opening = entry.get("opening_time") or "—"
            closing = entry.get("closing_time") or "—"
            status = "Present" if entry.get("opening_time") else "Absent"
            data_rows.append({
                "Sr No": str(sr),
                "Attendance ID": entry.get("attendance_id", ""),
                "Opening Time": opening,
                "Closing Time": closing,
                "Status": status
            })

        return header_info, data_rows
    except Exception as e:
        print(f"JSON read error: {e}")
        return {}, []


class UserReportWindow(tk.Toplevel):
    """Shows batch summary for a selected user."""

    def __init__(self, parent, drive_manager, lm, username, pin):
        super().__init__(parent)
        self.drive_manager = drive_manager
        self.lm = lm
        self.username = username
        self.pin = pin
        self.title(f"Report — {username}")
        self.geometry("750x500")
        self.resizable(True, True)
        try:
            self.grab_set()
        except Exception:
            pass

        self._build_ui()
        self.after(100, self._load_data)

    def _build_ui(self):
        # Header
        hdr = ttk.Frame(self)
        hdr.pack(fill=tk.X, padx=15, pady=(12, 5))
        ttk.Label(hdr, text=f"👤  {self.username}", font=("Arial", 14, "bold")).pack(side=tk.LEFT)
        ttk.Label(hdr, text=f"PIN: {self.pin}", font=("Arial", 10), foreground="gray").pack(side=tk.LEFT, padx=12)
        ttk.Button(hdr, text="↺  Refresh", command=self._load_data).pack(side=tk.RIGHT)

        # Loading label
        self.loading_lbl = ttk.Label(self, text="Loading data from Google Drive…", font=("Arial", 11), foreground="blue")
        self.loading_lbl.pack(pady=8)

        # Treeview
        cols = ("batch_id", "total", "success", "marked", "cut")
        frame = ttk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        sb = ttk.Scrollbar(frame)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", yscrollcommand=sb.set)
        sb.config(command=self.tree.yview)

        headings = {"batch_id": "Batch ID", "total": "Total Students", "success": "✓ Success",
                    "marked": "M Marked", "cut": "N Cut"}
        widths = {"batch_id": 150, "total": 110, "success": 100, "marked": 100, "cut": 80}
        for c in cols:
            self.tree.heading(c, text=headings[c])
            self.tree.column(c, width=widths[c], anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Detail button
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=15, pady=(0, 12))
        ttk.Button(btn_frame, text="📅  View Detail", command=self._open_detail).pack(side=tk.LEFT)

    def _load_data(self):
        self.loading_lbl.config(text="Loading data from Google Drive…", foreground="blue")
        self.tree.delete(*self.tree.get_children())
        threading.Thread(target=self._fetch_and_display, daemon=True).start()

    def _fetch_and_display(self):
        """Download all available data for this user and compute batch summaries."""
        dm = self.drive_manager
        batch_stats = {}  # batch_id → {total, success, marked, cut}

        try:
            dates = dm.get_cached_dates_for_user(self.username)
            # Also check Drive for any uncached dates
            drive_dates = dm.list_dates_for_user(self.username)
            all_dates = sorted(set(dates) | set(d for d, _ in drive_dates))

            for date_str in all_dates:
                batches_on_drive = dm.list_batches_for_date(self.username, date_str)
                for batch_id, file_id in batches_on_drive:
                    cache_path = dm.get_cached_json_path(self.username, date_str, batch_id)
                    if not os.path.exists(cache_path):
                        dm.download_json(file_id, cache_path)
                    if os.path.exists(cache_path):
                        hdr, rows = _read_json_to_rows(cache_path)
                        if batch_id not in batch_stats:
                            total = int(hdr.get("Total Students", len(rows)))
                            batch_stats[batch_id] = {"total": total, "success": 0, "marked": 0, "cut": 0}
                        for row in rows:
                            status = row.get("Status", "")
                            if status == "Present":
                                batch_stats[batch_id]["success"] += 1
                            elif status == "Absent":
                                batch_stats[batch_id]["cut"] += 1
        except Exception as e:
            self.after(0, lambda: self.loading_lbl.config(text=f"Error: {e}", foreground="red"))
            return

        def _update_ui():
            self.tree.delete(*self.tree.get_children())
            for bid, s in sorted(batch_stats.items()):
                self.tree.insert("", tk.END, iid=bid, values=(
                    bid, s["total"], s["success"], s["marked"], s["cut"]
                ))
            self.loading_lbl.config(
                text=f"Loaded {len(batch_stats)} batch(es)" if batch_stats else "No data found on Google Drive.",
                foreground="green" if batch_stats else "gray"
            )
        self.after(0, _update_ui)

    def _open_detail(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Batch", "Please select a batch to view details.", parent=self)
            return
        batch_id = sel[0]
        BatchDetailWindow(self, self.drive_manager, self.lm, self.username, self.pin, batch_id)


class BatchDetailWindow(tk.Toplevel):
    """Calendar date picker + daily sheet view + XLSX download + End Batch."""

    def __init__(self, parent, drive_manager, lm, username, pin, batch_id):
        super().__init__(parent)
        self.drive_manager = drive_manager
        self.lm = lm
        self.username = username
        self.pin = pin
        self.batch_id = batch_id
        self.title(f"Detail — {username} / Batch {batch_id}")
        self.geometry("860x640")
        self.resizable(True, True)
        self._current_rows = []
        self._current_header = {}
        self._current_date = None
        self._available_dates = set()
        self._build_ui()
        self.after(100, self._load_available_dates)

    def _build_ui(self):
        # Top: Calendar on left, info on right
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X, padx=15, pady=(10, 5))

        # Calendar
        cal_frame = ttk.LabelFrame(top_frame, text="Select Date")
        cal_frame.pack(side=tk.LEFT, padx=(0, 15))

        try:
            from tkcalendar import Calendar
            self.cal = Calendar(cal_frame, selectmode="day", date_pattern="yyyy-mm-dd",
                                font=("Arial", 9))
            self.cal.pack(padx=8, pady=8)
            self.cal.bind("<<CalendarSelected>>", self._on_date_selected)
        except ImportError:
            # Fallback: simple Entry
            ttk.Label(cal_frame, text="Date (YYYY-MM-DD):").pack(padx=5, pady=5)
            self.date_var = tk.StringVar()
            ttk.Entry(cal_frame, textvariable=self.date_var, width=15).pack(padx=5)
            ttk.Button(cal_frame, text="Load", command=self._on_date_entry).pack(pady=5)
            self.cal = None

        # Batch info panel (right of calendar)
        info_frame = ttk.LabelFrame(top_frame, text="Batch Information")
        info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.lbl_batch_id   = ttk.Label(info_frame, text="Batch ID: —",        font=("Arial", 11, "bold"))
        self.lbl_date       = ttk.Label(info_frame, text="Date: —",             font=("Arial", 10))
        self.lbl_in_time    = ttk.Label(info_frame, text="In Time: —",          font=("Arial", 10))
        self.lbl_out_time   = ttk.Label(info_frame, text="Out Time: —",         font=("Arial", 10))
        self.lbl_total      = ttk.Label(info_frame, text="Total Students: —",   font=("Arial", 10))
        self.lbl_status     = ttk.Label(info_frame, text="",                    font=("Arial", 10), foreground="blue")

        for lbl in (self.lbl_batch_id, self.lbl_date, self.lbl_in_time,
                    self.lbl_out_time, self.lbl_total, self.lbl_status):
            lbl.pack(anchor=tk.W, padx=12, pady=3)

        ttk.Button(info_frame, text="⬇  Download XLSX", command=self._download_xlsx).pack(
            anchor=tk.W, padx=12, pady=(8, 4))
            
        ttk.Button(info_frame, text="🛑  End Batch", command=self._end_batch).pack(
            anchor=tk.W, padx=12, pady=(0, 4))

        # Sheet treeview
        sheet_frame = ttk.LabelFrame(self, text="Attendance Sheet")
        sheet_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 12))

        cols = ("sr", "att_id", "opening", "closing", "status")
        sb = ttk.Scrollbar(sheet_frame)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.sheet_tree = ttk.Treeview(sheet_frame, columns=cols, show="headings", yscrollcommand=sb.set)
        sb.config(command=self.sheet_tree.yview)

        heads = {"sr": "Sr No", "att_id": "Attendance ID", "opening": "Opening Time",
                 "closing": "Closing Time", "status": "Status"}
        widths = {"sr": 60, "att_id": 180, "opening": 130, "closing": 130, "status": 90}
        for c in cols:
            self.sheet_tree.heading(c, text=heads[c])
            self.sheet_tree.column(c, width=widths[c], anchor=tk.CENTER)

        # Tag colors
        self.sheet_tree.tag_configure("absent", foreground="red")
        self.sheet_tree.tag_configure("present", foreground="green")

        self.sheet_tree.pack(fill=tk.BOTH, expand=True)

    def _load_available_dates(self):
        """Fetch all dates that have data for this batch from cache/Drive."""
        def _worker():
            dm = self.drive_manager
            dates = set()
            cached = dm.get_cached_dates_for_user(self.username)
            for d in cached:
                batches = dm.get_cached_batches_for_date(self.username, d)
                if self.batch_id in batches:
                    dates.add(d)
            # Also check Drive
            drive_dates = dm.list_dates_for_user(self.username)
            for d_str, _ in drive_dates:
                batches = dm.list_batches_for_date(self.username, d_str)
                for bid, _ in batches:
                    if bid == self.batch_id:
                        dates.add(d_str)
            self.after(0, lambda: self._apply_available_dates(dates))
        threading.Thread(target=_worker, daemon=True).start()

    def _apply_available_dates(self, dates):
        self._available_dates = dates

        if not dates:
            self.lbl_status.config(text="No data available for this batch.", foreground="gray")
            return

        # Set calendar min date if possible
        sorted_dates = sorted(dates)
        earliest = sorted_dates[0]

        if hasattr(self, 'cal') and self.cal is not None:
            try:
                from tkcalendar import Calendar
                # Gray out dates not in our set — tkcalendar doesn't natively support disabling,
                # but we validate on selection instead.
                min_dt = datetime.strptime(earliest, "%Y-%m-%d").date()
                self.cal.config(mindate=min_dt)
                # Highlight available dates using events
                for d in dates:
                    dt = datetime.strptime(d, "%Y-%m-%d")
                    self.cal.calevent_create(dt, "Data available", "available")
                self.cal.tag_config("available", background="#c8f7c5", foreground="black")
            except Exception:
                pass

        self.lbl_status.config(
            text=f"{len(dates)} date(s) with data. Select a date to view.",
            foreground="blue"
        )

    def _on_date_selected(self, event=None):
        if hasattr(self, 'cal') and self.cal:
            selected = self.cal.get_date()
        else:
            return
        if selected not in self._available_dates:
            self.lbl_status.config(text=f"No data for {selected}.", foreground="orange")
            self.sheet_tree.delete(*self.sheet_tree.get_children())
            return
        self._load_sheet_for_date(selected)

    def _on_date_entry(self):
        if hasattr(self, 'date_var'):
            d = self.date_var.get().strip()
            if d:
                self._load_sheet_for_date(d)

    def _load_sheet_for_date(self, date_str):
        self._current_date = date_str
        self.lbl_status.config(text=f"Loading {date_str}…", foreground="blue")
        self.sheet_tree.delete(*self.sheet_tree.get_children())

        def _worker():
            dm = self.drive_manager
            cache_path = dm.get_cached_json_path(self.username, date_str, self.batch_id)
            if not os.path.exists(cache_path):
                ok, result = dm.download_batch_json(self.username, date_str, self.batch_id)
                if not ok:
                    self.after(0, lambda: self.lbl_status.config(
                        text=f"Failed: {result}", foreground="red"))
                    return
                cache_path = result

            hdr, rows = _read_json_to_rows(cache_path)
            self.after(0, lambda h=hdr, r=rows: self._populate_sheet(h, r, date_str))
        threading.Thread(target=_worker, daemon=True).start()

    def _populate_sheet(self, hdr, rows, date_str):
        self._current_header = hdr
        self._current_rows = rows
        self._current_date = date_str

        self.lbl_batch_id.config(text=f"Batch ID: {hdr.get('Batch ID', self.batch_id)}")
        self.lbl_date.config(text=f"Date: {hdr.get('Date', date_str)}")
        self.lbl_in_time.config(text=f"In Time: {hdr.get('In Time', '—')}")
        self.lbl_out_time.config(text=f"Out Time: {hdr.get('Out Time', '—')}")
        self.lbl_total.config(text=f"Total Students: {hdr.get('Total Students', len(rows))}")

        self.sheet_tree.delete(*self.sheet_tree.get_children())
        for row in rows:
            status = row.get("Status", "Absent")
            tag = "present" if status == "Present" else "absent"
            self.sheet_tree.insert("", tk.END, tags=(tag,), values=(
                row.get("Sr No", ""),
                row.get("Attendance ID", ""),
                row.get("Opening Time", "—"),
                row.get("Closing Time", "—"),
                status
            ))

        present = sum(1 for r in rows if r.get("Status") == "Present")
        absent = len(rows) - present
        self.lbl_status.config(
            text=f"Present: {present}   Absent: {absent}   Total: {len(rows)}",
            foreground="green"
        )

    def _download_xlsx(self):
        if not self._current_rows and not self._current_header:
            messagebox.showwarning("No Data", "Please select a date first.", parent=self)
            return

        default_name = f"{self.batch_id}_{self._current_date or 'export'}.xlsx"
        save_path = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=default_name,
            title="Save Attendance Report"
        )
        if not save_path:
            return

        try:
            _write_xlsx(save_path, self._current_header, self._current_rows,
                        self.batch_id, self._current_date or "")
            messagebox.showinfo("Downloaded", f"File saved to:\n{save_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}", parent=self)

    def _end_batch(self):
        if messagebox.askyesno("End Batch", f"Stop monitoring batch {self.batch_id} for user {self.username}?\nThis will hide it from their IR Attendance software.", parent=self):
            success, msg = self.lm.end_batch(self.pin, self.batch_id)
            if success:
                messagebox.showinfo("Success", f"Batch {self.batch_id} ended.", parent=self)
            else:
                messagebox.showerror("Error", f"Failed to end batch: {msg}", parent=self)


def _write_xlsx(path, hdr, rows, batch_id, date_str):
    """Write attendance data to an XLSX file with proper header."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Styles
    bold14 = Font(bold=True, size=14)
    bold11 = Font(bold=True, size=11)
    bold10 = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    absent_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"IR Attendance Report — Batch {batch_id}"
    ws["A1"].font = bold14
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 24

    # Row 2: Meta info
    meta_cols = [
        ("A2", f"Batch ID: {hdr.get('Batch ID', batch_id)}"),
        ("B2", f"Date: {hdr.get('Date', date_str)}"),
        ("C2", f"In Time: {hdr.get('In Time', '—')}"),
        ("D2", f"Out Time: {hdr.get('Out Time', '—')}"),
        ("E2", f"Total Students: {hdr.get('Total Students', len(rows))}"),
    ]
    for cell_ref, text in meta_cols:
        ws[cell_ref] = text
        ws[cell_ref].font = bold11
    ws.row_dimensions[2].height = 18

    # Row 3: blank
    ws.row_dimensions[3].height = 8

    # Row 4: Column headers
    col_headers = ["Sr No", "Attendance ID", "Opening Time", "Closing Time", "Status"]
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 18

    # Data rows
    for i, row in enumerate(rows, 5):
        status = row.get("Status", "Absent")
        fill = present_fill if status == "Present" else absent_fill
        values = [
            row.get("Sr No", i - 4),
            row.get("Attendance ID", ""),
            row.get("Opening Time", "—"),
            row.get("Closing Time", "—"),
            status
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = center
            cell.border = border

    # Column widths
    for col, width in zip("ABCDE", [10, 20, 16, 16, 12]):
        ws.column_dimensions[col].width = width

    wb.save(path)
